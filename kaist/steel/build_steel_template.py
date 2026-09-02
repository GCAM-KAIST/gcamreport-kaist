# -*- coding: utf-8 -*-
"""
build_steel_template.py

Fill the KMIP2026 steel-sector DB template from GCAM ModelInterface query
results (South Korea).

Usage (from the repo root; --dir is the data folder):
  python -X utf8 kaist/steel/build_steel_template.py \
      --dir kmip/KMIP25_6Scenarios_output/iron_report              # S1
  python -X utf8 kaist/steel/build_steel_template.py \
      --dir kmip/KMIP25_6Scenarios_output/iron_report \
      --tag kaist9_nz --input query_results_kaist9.xlsx \
      --xml kaist/input/Korea/iron_steel_no_h2.xml --base-year 2021
  Running inside the data folder without --dir still works.

Inputs  (in --dir unless a path is given):
  - query_results*.xlsx                          : raw GCAM query results
  - (붙임 1) KMIP2026_DB_template_steel_v0.xlsx : template to fill
  - (붙임 2-1) ... (참고용).xlsx                : KMIP2025 MT results (ratios)
  - iron_steel XML of the GCAM run              : coal coefficients by tech
    (--xml is resolved from the current folder; default = the repo's
    gcam_input/gcamdata/xml/iron_steel.xml)

Outputs (in --dir):
  - KMIP2026_steel_GC_{tag}.xlsx   : filled template (format/row order kept;
                                     if it already exists, note column M is
                                     preserved and only values are rewritten)
  - steel_intermediates_{tag}.xlsx : traceable intermediate tables

Method summary (details: steel_report_memo.md):
  - Final Energy by fuel from sector inputs (EJ -> ktoe); scrap and the
    policy dummy inputs are excluded.
  - Biomass|Liquids = liquids consumption x national refining biomass share,
    removed from Oil; Biomass|Solids = delivered biomass.
  - Coal Fuel/Feedstock: BF-family coal x MT-2020 feedstock share; all other
    coal (EAF-DRI heating) = Fuel. BF-family coal = sector coal minus
    non-BF coal reconstructed from production x XML coefficient.
  - CO2 (energy) basis = "no bio" query when available (biogenic CO2
    neutral, IAMC convention, same as gcamreport); by-tech query otherwise.
  - Industrial Process CO2 = Coal|Feedstock (ktoe) x MT-2020
    (process CO2 / coal feedstock) factor; subtracted from the energy row.
  - CH4/N2O: AR6 GWP100 (27.2 / 273), same as kaist step4.
No estimation beyond the rules above; missing data -> blank + note.
"""

import argparse
import os
import xml.etree.ElementTree as ET

import pandas as pd
import openpyxl

# ----------------------------------------------------------------------------
# Arguments
# ----------------------------------------------------------------------------
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

ap = argparse.ArgumentParser()
ap.add_argument("--dir", default=".",
                help="data folder holding query_results*.xlsx, the template and "
                     "the MT reference; outputs are written there (default: .)")
ap.add_argument("--input", default=None,
                help="query results workbook, relative to --dir "
                     "(default: query_results_{tag}.xlsx, else query_results.xlsx)")
ap.add_argument("--tag", default="S1", help="output file suffix")
ap.add_argument("--xml", default=os.path.join(REPO_ROOT, "gcam_input", "gcamdata",
                                              "xml", "iron_steel.xml"),
                help="iron_steel XML of the GCAM run (coal coefficients); "
                     "relative paths are resolved from the current folder")
ap.add_argument("--scenario", default=None,
                help="scenario name prefix; auto-detected if the file has one")
ap.add_argument("--no-process-split", action="store_true",
                help="leave the Industrial Process row blank")
ap.add_argument("--extra-years", type=int, nargs="*", default=[],
                help="additional years to compute (not written to the "
                     "template; saved in the intermediates), e.g. 2015")
ap.add_argument("--base-year", type=int, default=2020,
                help="first reported year (2021 for GCAM 9.x runs); replaces "
                     "the template's 2020 column header")
args = ap.parse_args()

# --xml is resolved before moving into the data folder; everything else
# (inputs, template, reference, outputs) is relative to --dir.
IRON_STEEL_XML = os.path.abspath(args.xml)
if not os.path.isdir(args.dir):
    raise SystemExit(f"--dir not found: {args.dir}")
os.chdir(args.dir)

TAG = args.tag
QUERY_FILE = args.input
if QUERY_FILE is None:
    QUERY_FILE = f"query_results_{TAG}.xlsx"
    if not os.path.exists(QUERY_FILE):
        QUERY_FILE = "query_results.xlsx"
TEMPLATE_FILE = "(붙임 1) KMIP2026_DB_template_steel_v0.xlsx"
REFERENCE_FILE = ("(붙임 2-1) KMIP2025 상향식 모형 분석 결과_장기감축경로_"
                  "MT+GC+MS (참고용).xlsx")
OUT_TEMPLATE = f"KMIP2026_steel_GC_{TAG}.xlsx"
OUT_INTERMEDIATE = f"steel_intermediates_{TAG}.xlsx"
FILL_PROCESS_SPLIT = not args.no_process_split

REGION = "South Korea"
SECTOR = "iron and steel"
TEMPLATE_YEARS = [args.base_year, 2025, 2030, 2035, 2040, 2045, 2050]
YEARS = sorted(set(args.extra_years)) + TEMPLATE_YEARS   # all computed years

# Unit conversion constants
# 1 ktoe = 41.868 TJ (IEA)  ->  1 EJ = 1e6 / 41.868 = 23884.6 ktoe
EJ_TO_KTOE = 23884.6
# GCAM reports carbon mass (MTC). CO2 = C x 44/12
MTC_TO_MTCO2 = 44.0 / 12.0
# GWP100 AR6 -- same factors as kaist/step4_fill_template.R unit table
GWP_CH4 = 27.2
GWP_N2O = 273.0

# GCAM input name -> KMIP fuel
FUEL_MAP = {
    "delivered coal": "Coal",
    "wholesale gas": "Gas",
    "refined liquids industrial": "Oil",
    "elect_td_ind": "Electricity",
    "H2 industrial": "Hydrogen",
    "delivered biomass": "Biomass",
}
# Not energy flows: scrap (Mt of steel scrap), policy-constraint dummies
EXCLUDED_INPUTS = ["scrap", "irnstl-ceiling", "irnstl_ceiling_EAF"]

SHEETS = {
    "inputs": "iron and steel inputs by tech (",
    "prod_region": "iron and steel production by re",
    "prod_tech": "iron and steel production by te",
    "co2": "CO2 emissions by tech (excludin",
    "co2_nobio": "CO2 emissions by tech (no bio) ",
    "seq": "CO2 sequestration by tech",
    "nonco2": "nonCO2 emissions by sector (exc",
    "refliq": "refined liquids production by s",
}


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def read_sheet(xl, key, required=True):
    """Read a query sheet by short key. Row 1 is normally the query title
    and row 2 the header; a sheet whose A1 is 'scenario' has no title row.
    Returns None for a missing optional sheet."""
    sheet = SHEETS[key]
    if sheet not in xl.sheet_names:
        if required:
            raise ValueError(f"Required sheet '{sheet}' not in {QUERY_FILE}")
        print(f"[MISSING] optional sheet '{sheet}' not in {QUERY_FILE}")
        return None
    peek = pd.read_excel(xl, sheet, header=None, nrows=1)
    skip = 0 if str(peek.iloc[0, 0]).strip() == "scenario" else 1
    df = pd.read_excel(xl, sheet, skiprows=skip)
    if "scenario" not in df.columns:          # e.g. "The query returned no results"
        msg = f"sheet '{sheet}' has no data table"
        if required:
            raise ValueError(msg)
        print(f"[MISSING] {msg} -> treated as absent")
        return None
    present = sorted(df["scenario"].astype(str).str.split(",").str[0].unique())
    df = df[df["region"] == REGION]
    df = df[df["scenario"].astype(str).str.startswith(SCENARIO_PREFIX)]
    if len(df) == 0:
        msg = (f"sheet '{sheet}' has no rows for scenario '{SCENARIO_PREFIX}' "
               f"(scenarios present: {present})")
        if required:
            raise ValueError(msg)
        print(f"[MISSING] {msg} -> treated as absent")
        return None
    return df


def year_values(df_rows, factor=1.0):
    return {y: float(df_rows[y].sum()) * factor for y in YEARS}


def check_units(df_rows, expected, label):
    units = set(df_rows["Units"].astype(str).str.strip())
    if units and units != {expected}:
        raise ValueError(f"[UNITS] {label}: expected {{{expected}}}, got {units}")
    print(f"[UNITS OK] {label}: {units or '{(no rows)}'} == {expected}")


# ----------------------------------------------------------------------------
# 1. Load query results
# ----------------------------------------------------------------------------
xl = pd.ExcelFile(QUERY_FILE)

# scenario prefix: given, or auto-detected from the first sheet
if args.scenario:
    SCENARIO_PREFIX = args.scenario
else:
    peek = pd.read_excel(xl, SHEETS["inputs"], header=None, nrows=3)
    col0 = peek.iloc[:, 0].astype(str)
    scen_cell = next(c for c in col0 if "," in c or c.startswith("S"))
    SCENARIO_PREFIX = scen_cell.split(",")[0]
print(f"Input: {QUERY_FILE} | scenario prefix: '{SCENARIO_PREFIX}' | tag: {TAG}")

inputs = read_sheet(xl, "inputs")
prod_region = read_sheet(xl, "prod_region")
prod_tech = read_sheet(xl, "prod_tech")
co2 = read_sheet(xl, "co2")
co2_nobio = read_sheet(xl, "co2_nobio", required=False)
seq = read_sheet(xl, "seq")
nonco2 = read_sheet(xl, "nonco2")
refliq = read_sheet(xl, "refliq", required=False)

# Years with model output. GCAM 9.x has base year 2021, so its 2020 column
# is all zero -> reported blank, not zero.
ist_prod = prod_region[prod_region["sector"] == SECTOR]
check_units(ist_prod, "Mt", "steel production")
steel_prod = year_values(ist_prod)
ACTIVE = {y: steel_prod[y] > 0 for y in YEARS}
inactive = [y for y in YEARS if not ACTIVE[y]]
if inactive:
    print(f"[YEARS] no model output for {inactive} -> left blank in template")

# ----------------------------------------------------------------------------
# 2. Final Energy by fuel (EJ -> ktoe)
# ----------------------------------------------------------------------------
ist_inputs = inputs[inputs["sector"] == SECTOR]
print("\n=== Final Energy input handling ===")
exclusion_log = []
for _, row in ist_inputs.iterrows():
    name = row["input"]
    if name in FUEL_MAP:
        print(f"[INCLUDE] '{name}' -> {FUEL_MAP[name]} (Units={row['Units']})")
    elif name in EXCLUDED_INPUTS:
        print(f"[EXCLUDE] '{name}' (Units={row['Units']}) -- not an energy flow")
        exclusion_log.append({"input": name, "Units": row["Units"],
                              **{y: row[y] for y in YEARS}})
    else:
        raise ValueError(f"Unexpected input '{name}' -- classify before rerunning.")

fe_ej, fe_rows = {}, {}
for gcam_name, fuel in FUEL_MAP.items():
    rows = ist_inputs[ist_inputs["input"] == gcam_name]
    check_units(rows, "EJ", f"FE {fuel} ({gcam_name})")
    fe_ej[fuel] = year_values(rows)
    fe_rows[fuel] = year_values(rows, EJ_TO_KTOE)

# ----------------------------------------------------------------------------
# 2b. Biomass Solids / Liquids split (national refining biomass share).
#     All liquid fuel in GCAM comes from one blended national refining pool,
#     so the bio share of any sector's liquids = national share. The bio
#     part is moved out of Oil so fuels still sum to the total.
# ----------------------------------------------------------------------------
print("\n=== Biomass Liquids split ===")
if refliq is not None:
    check_units(refliq, "EJ", "refined liquids production")
    # GCAM 7: one 'refining' sector with subsector 'biomass liquids';
    # GCAM 9: separate sectors ('oil refining', 'biomass liquids', ...),
    # so match the name in either column.
    is_bio = (refliq.get("subsector", pd.Series("", index=refliq.index)) == "biomass liquids") | \
             (refliq["sector"] == "biomass liquids")
    bio_share = {}
    for y in YEARS:
        tot = float(refliq[y].sum())
        bio = float(refliq[is_bio][y].sum())
        bio_share[y] = bio / tot if tot > 0 else 0.0
    print("[SPLIT] refining biomass share = "
          + ", ".join(f"{y}:{bio_share[y]:.4f}" for y in YEARS))
    fe_bio_solids = dict(fe_rows["Biomass"])
    fe_bio_liquids = {y: fe_rows["Oil"][y] * bio_share[y] for y in YEARS}
    fe_rows["Oil"] = {y: fe_rows["Oil"][y] * (1 - bio_share[y]) for y in YEARS}
    fe_rows["Biomass"] = {y: fe_bio_solids[y] + fe_bio_liquids[y] for y in YEARS}
    BIO_SPLIT_OK = True
else:
    print("[SKIP] no 'refined liquids production by subsector' sheet -> "
          "Oil left unsplit; Biomass|Solids/Liquids rows held")
    bio_share = {y: None for y in YEARS}
    fe_bio_solids = dict(fe_rows["Biomass"])
    fe_bio_liquids = None
    BIO_SPLIT_OK = False
fe_total = {y: sum(fe_rows[f][y] for f in fe_rows) for y in YEARS}

# ----------------------------------------------------------------------------
# 2c. Coal Fuel / Feedstock split.
#     BF-family coal = sector coal - non-BF coal, where non-BF coal (EAF-DRI
#     routes, heating only) = production x delivered-coal coefficient from
#     the run's iron_steel XML (latest defined period carried forward; a CCS
#     tech without its own coefficient uses its non-CCS sibling's).
#     BF-family coal x MT-2020 feedstock share -> Feedstock; rest -> Fuel.
#     When every producing tech has a coefficient (GCAM 7 XML), the full
#     reconstruction is also checked against the queried sector coal (1%).
# ----------------------------------------------------------------------------
print("\n=== Coal Fuel/Feedstock split ===")
coef_by_tech = {}      # tech -> {year: coal coef}
techs_with_inputs = set()   # techs whose energy inputs are listed in the XML
for reg in ET.parse(IRON_STEEL_XML).getroot().iter("region"):
    if reg.get("name") != REGION:
        continue
    for sec in reg.iter("supplysector"):
        if sec.get("name") != SECTOR:
            continue
        for sub in sec.iter("subsector"):
            for tech in list(sub.iter("technology")) + list(sub.iter("stub-technology")):
                for per in tech.iter("period"):
                    for mi in per.iter("minicam-energy-input"):
                        techs_with_inputs.add(tech.get("name"))
                        if mi.get("name") == "delivered coal" and mi.find("coefficient") is not None:
                            coef_by_tech.setdefault(tech.get("name"), {})[int(per.get("year"))] = \
                                float(mi.find("coefficient").text)


def coal_coef(tech, year):
    """Coal coefficient for tech/year: exact year, else latest earlier
    period, else the non-CCS sibling's. A tech whose inputs are listed
    without coal uses 0. None only if the XML says nothing about it."""
    for name in (tech, tech.replace(" CCS", "")):
        cs = coef_by_tech.get(name)
        if cs:
            yrs = [y for y in cs if y <= year]
            if yrs:
                return cs[max(yrs)], name
        if name in techs_with_inputs:
            return 0.0, f"{name} (no coal input)"
    return None, None


coal_rows, coef_log = [], []
for _, p in prod_tech.iterrows():
    tech, sub = p["technology"], str(p["subsector"])
    row = {"family": "BLASTFUR" if sub.startswith("BLASTFUR") else "EAF",
           "subsector": sub, "technology": tech}
    known = True
    for y in YEARS:
        c, src = coal_coef(tech, y)
        row[y] = float(p[y]) * c if c is not None else None
        if c is None and float(p[y]) > 0:
            known = False
    row["coef_known"] = known
    c2050, src = coal_coef(tech, 2050)
    coef_log.append({"technology": tech, "coef_2050_EJ_per_Mt": c2050, "coef_source": src})
    coal_rows.append(row)
coal_by_tech = pd.DataFrame(coal_rows)
for r in coef_log:
    print(f"[COEF] {r['technology']:24} coal coef(2050)={r['coef_2050_EJ_per_Mt']} "
          f"(from '{r['coef_source']}')")

nonbf = coal_by_tech[coal_by_tech["family"] != "BLASTFUR"]
if not nonbf["coef_known"].all():
    raise ValueError("Non-BF tech with production but no coal coefficient: "
                     f"{nonbf.loc[~nonbf['coef_known'], 'technology'].tolist()}")
nonbf_coal = {y: float(nonbf[y].fillna(0).sum()) for y in YEARS}
bf_share = {y: (1 - nonbf_coal[y] / fe_ej["Coal"][y]) if fe_ej["Coal"][y] > 0 else 0.0
            for y in YEARS}

# Strict 1% check only when every coefficient is the tech's own (GCAM 7
# XML). Borrowed sibling coefficients (GCAM 9 XML lists only calibrated
# techs) make the reconstruction approximate -> report the gap instead.
exact = all(r["coef_source"] in (r["technology"], f"{r['technology']} (no coal input)")
            for r in coef_log)
gaps = {y: (float(coal_by_tech[y].sum()) - fe_ej["Coal"][y]) / fe_ej["Coal"][y]
        for y in YEARS if fe_ej["Coal"][y] > 1e-9}
if exact:
    bad = {y: g for y, g in gaps.items() if abs(g) > 0.01}
    if bad:
        raise ValueError(f"[COAL CHECK] reconstruction off by >1%: {bad}")
    print("[COAL CHECK] full reconstruction matches queried sector coal within 1%")
else:
    print("[COAL CHECK] approximate (borrowed coefficients); reconstruction gap "
          "vs queried coal = " + ", ".join(f"{y}:{g:+.1%}" for y, g in gaps.items())
          + " -- BF coal taken as sector coal minus non-BF coal")
print("[COAL] BF-family share of coal = "
      + ", ".join(f"{y}:{bf_share[y]:.4f}" for y in YEARS))

# MT 2020 ratios from the KMIP2025 reference (MT 2020 coal assumed all-BF)
ref = pd.read_excel(REFERENCE_FILE, "data")
mt = ref[(ref["Model"] == "MT") & (ref["Scenario"].astype(str).str.strip() == "S1")]
MT_FE = "Final Energy|Industry|Manufacturing|Iron and Steel|"


def mt_val(variable, year):
    rows = mt[mt["Variable"] == variable]
    if len(rows) != 1:
        raise ValueError(f"Reference: expected 1 MT S1 row for '{variable}', got {len(rows)}")
    return float(rows.iloc[0][str(year)])


mt_fuel20 = mt_val(MT_FE + "Coal|Fuel", 2020)
mt_feed20 = mt_val(MT_FE + "Coal|Feedstock", 2020)
FS2020 = mt_feed20 / (mt_fuel20 + mt_feed20)
print(f"[COAL] MT 2020 feedstock share = {FS2020:.4f}")

fe_split = {"Coal": {}}
fe_split["Coal"]["Feedstock"] = {y: fe_rows["Coal"][y] * bf_share[y] * FS2020 for y in YEARS}
fe_split["Coal"]["Fuel"] = {y: fe_rows["Coal"][y] - fe_split["Coal"]["Feedstock"][y] for y in YEARS}
# other fuels follow MT's constant classification
for fuel, fs in (("Biomass", 0.0), ("Oil", 0.0), ("Gas", 0.0), ("Hydrogen", 1.0)):
    fe_split[fuel] = {"Feedstock": {y: fe_rows[fuel][y] * fs for y in YEARS},
                      "Fuel": {y: fe_rows[fuel][y] * (1 - fs) for y in YEARS}}

# ----------------------------------------------------------------------------
# 3. CO2 emissions (MTC -> Mt CO2)
#    Energy row basis: "no bio" total when available (biogenic CO2 treated
#    as neutral, IAMC convention and gcamreport behaviour); the by-tech query
#    (includes biomass-BF CO2) is kept for the breakdown table.
# ----------------------------------------------------------------------------
print("\n=== CO2 ===")
ist_co2 = co2[co2["sector"] == SECTOR]
check_units(ist_co2, "MTC", "CO2 by tech")
co2_by_tech = ist_co2[["subsector", "technology"] + YEARS].copy()
for y in YEARS:
    co2_by_tech[y] = co2_by_tech[y] * MTC_TO_MTCO2
co2_by_tech.insert(0, "family", co2_by_tech["subsector"].astype(str)
                   .apply(lambda s: "BLASTFUR" if s.startswith("BLASTFUR") else "EAF"))
co2_by_tech["Units"] = "Mt CO2/yr"
co2_bytech_total = {y: float(co2_by_tech[y].sum()) for y in YEARS}

if co2_nobio is not None:
    ist_nb = co2_nobio[co2_nobio["sector"] == SECTOR]
    check_units(ist_nb, "MTC", "CO2 no bio")
    co2_total = year_values(ist_nb, MTC_TO_MTCO2)
    CO2_BASIS = "no bio"
else:
    co2_total = dict(co2_bytech_total)
    CO2_BASIS = "by tech (no 'no bio' sheet available)"
print(f"[CO2] basis = {CO2_BASIS}; 2050 by-tech={co2_bytech_total[2050]:.3f}, "
      f"used={co2_total[2050]:.3f} Mt CO2")

# Industrial process CO2 = Coal|Feedstock (ktoe) x MT-2020 factor
mt_proc20 = mt_val("Emissions|GHGs|Non-Energy|Industrial Process|Manufacturing|Iron and Steel", 2020)
PROC_FACTOR = mt_proc20 / mt_feed20          # Mt CO2 per ktoe of coal feedstock
process_co2 = {y: fe_split["Coal"]["Feedstock"][y] * PROC_FACTOR for y in YEARS}
energy_co2 = {y: co2_total[y] - process_co2[y] for y in YEARS}
print(f"[PROCESS] factor = MT2020 process {mt_proc20:.3f} / coal feedstock "
      f"{mt_feed20:.1f} ktoe = {PROC_FACTOR:.5f} Mt CO2/ktoe; applied={FILL_PROCESS_SPLIT}")

# ----------------------------------------------------------------------------
# 4. CCS, CH4/N2O
# ----------------------------------------------------------------------------
ist_seq = seq[seq["sector"] == SECTOR]
check_units(ist_seq, "MTC", "CO2 sequestration")
seq_total = year_values(ist_seq, MTC_TO_MTCO2)
seq_by_tech = ist_seq[["subsector", "technology"] + YEARS].copy()
for y in YEARS:
    seq_by_tech[y] = seq_by_tech[y] * MTC_TO_MTCO2

ist_nonco2 = nonco2[nonco2["sector"] == SECTOR]
check_units(ist_nonco2, "Tg", "nonCO2")
ghg_co2eq = {}
for gas, gwp in (("CH4", GWP_CH4), ("N2O", GWP_N2O)):
    rows = ist_nonco2[ist_nonco2["GHG"].astype(str).str.match(rf"^{gas}(_|$)")]
    ghg_co2eq[gas] = year_values(rows, gwp)        # Tg == Mt

# ----------------------------------------------------------------------------
# 5. Fill the template
# ----------------------------------------------------------------------------
BASE_IS_OUTPUT = os.path.exists(OUT_TEMPLATE)
wb = openpyxl.load_workbook(OUT_TEMPLATE if BASE_IS_OUTPUT else TEMPLATE_FILE)
ws = wb["data"]
print(f"\nBase workbook: {'existing ' + OUT_TEMPLATE + ' (notes preserved)' if BASE_IS_OUTPUT else TEMPLATE_FILE}")

YEAR_COL = dict(zip(TEMPLATE_YEARS, "FGHIJKL"))
NOTE_COL = "M"
ws["F1"] = args.base_year          # header of the first year column
V = "Final Energy|Industry|Iron and Steel"
NA_LIQ = "보류: 'refined liquids production by subsector' 쿼리 없음 (Oil 미분리)"

actions, notes = {}, {}


def fill(var, values, note=None):
    actions[var] = ("fill", values)
    if note:
        notes[var] = note


def hold(var, note):
    actions[var] = ("hold", None)
    notes[var] = note


fill("Emissions|CO2|Energy|Demand|Industry|Iron and Steel",
     energy_co2 if FILL_PROCESS_SPLIT else co2_total,
     f"기준: {CO2_BASIS}" + (" ; 공정배출 차감" if FILL_PROCESS_SPLIT else ""))
fill("Emissions|CH4|Energy|Demand|Industry|Iron and Steel", ghg_co2eq["CH4"], "AR6 GWP100 27.2 (step4와 동일)")
fill("Emissions|N2O|Energy|Demand|Industry|Iron and Steel", ghg_co2eq["N2O"], "AR6 GWP100 273 (step4와 동일)")
fill("Carbon Sequestration|CCS|Energy|Demand|Industry|Iron and Steel", seq_total)
if FILL_PROCESS_SPLIT:
    fill("Emissions|GHGs|Non-Energy|Industrial Process|Iron and Steel", process_co2,
         f"Coal|Feedstock(ktoe) × {PROC_FACTOR:.5f} MtCO2/ktoe (MT2020 공정배출/feedstock 비율)")
else:
    hold("Emissions|GHGs|Non-Energy|Industrial Process|Iron and Steel", "보류: 에너지 배출에 포함")
fill(V, fe_total)
for fuel in ("Coal", "Gas", "Oil", "Electricity", "Hydrogen", "Biomass"):
    fill(f"{V}|{fuel}", fe_rows[fuel],
         "refined liquids × (1−bio_share)" if fuel == "Oil" and BIO_SPLIT_OK else
         (NA_LIQ if fuel == "Oil" else None))
for fuel in ("Coal", "Biomass", "Oil", "Gas", "Hydrogen"):
    for kind in ("Fuel", "Feedstock"):
        fill(f"{V}|{fuel}|{kind}", fe_split[fuel][kind],
             "BF계열 coal × MT2020 feedstock 비율, 나머지 Fuel" if fuel == "Coal"
             else "H2 전량 Feedstock, 그 외 전량 Fuel (MT 분류)")
if BIO_SPLIT_OK:
    for s in ("biomass|Solids", "Biomass|Solids"):
        fill(f"{V}|{s}", fe_bio_solids, "delivered biomass")
    fill(f"{V}|Biomass|Liquids", fe_bio_liquids, "액체연료 소비 × 국가 정제 바이오 비중 (Oil에서 차감)")
else:
    for s in ("biomass|Solids", "Biomass|Solids"):
        fill(f"{V}|{s}", fe_bio_solids, "delivered biomass")
    hold(f"{V}|Biomass|Liquids", NA_LIQ)
for s, why in (("Heat", "district heat input 없음"), ("Solar", "solar input 없음"),
               ("Solar|PV", "상동"), ("Solar|Solarthermal", "상동"), ("Waste", "waste input 없음")):
    hold(f"{V}|{s}", f"NR: GCAM iron and steel에 {why}")
fill("Production|Iron and Steel|Steel", steel_prod)
hold("Production|Iron and Steel|Iron", "NR: GCAM 별도 output 없음 (MT도 0 제출)")
hold("Production|Iron and Steel|Iron Ore", "NR: GCAM 별도 output 없음 (MT도 0 제출)")

print("\n=== Template fill ===")
seen = set()
for r in range(2, ws.max_row + 1):
    var = ws[f"D{r}"].value
    if var is None:
        continue
    ws[f"A{r}"] = "GC"
    if var not in actions:
        raise ValueError(f"Template row {r}: unmapped variable '{var}'")
    seen.add(var)
    kind, payload = actions[var]
    for y in TEMPLATE_YEARS:
        ws[f"{YEAR_COL[y]}{r}"] = payload[y] if (kind == "fill" and ACTIVE[y]) else None
    if not BASE_IS_OUTPUT and var in notes:
        cell = ws[f"{NOTE_COL}{r}"]
        cell.value = f"{cell.value} | {notes[var]}" if cell.value else notes[var]
    print(f"[{kind.upper():4}] row {r}: {var}")

missing = set(actions) - seen - {f"{V}|biomass|Solids", f"{V}|Biomass|Solids"}
if missing:
    raise ValueError(f"Not found in template: {missing}")
wb.save(OUT_TEMPLATE)
print(f"Saved {OUT_TEMPLATE}")

# ----------------------------------------------------------------------------
# 6. Intermediates
# ----------------------------------------------------------------------------
def table(rows):
    return pd.DataFrame(rows)


summary = table([
    {"item": "CO2 by-tech total (incl. biomass BF)", **co2_bytech_total},
    {"item": f"CO2 used ({CO2_BASIS})", **co2_total},
    {"item": f"Industrial Process = Coal|Feedstock x {PROC_FACTOR:.5f}", **process_co2},
    {"item": "Energy|Demand CO2 = used - process", **energy_co2},
    {"item": "MT2020 coal feedstock share", **{y: FS2020 for y in YEARS}},
    {"item": "BF-family share of sector coal", **bf_share},
    {"item": "refining biomass share", **bio_share},
    {"item": "steel production (Mt)", **steel_prod},
])
fe_table = table(
    [{"variable": f, "Units": "ktoe/yr", **fe_rows[f]} for f in fe_rows]
    + [{"variable": "Biomass|Solids", "Units": "ktoe/yr", **fe_bio_solids}]
    + ([{"variable": "Biomass|Liquids", "Units": "ktoe/yr", **fe_bio_liquids}] if BIO_SPLIT_OK else [])
    + [{"variable": f"{f}|{k}", "Units": "ktoe/yr", **fe_split[f][k]}
       for f in fe_split for k in ("Fuel", "Feedstock")]
    + [{"variable": "TOTAL", "Units": "ktoe/yr", **fe_total}])

# every template variable x every computed year (incl. --extra-years)
template_long = table([
    {"variable": var, **{y: (payload[y] if (kind == "fill" and ACTIVE[y]) else None) for y in YEARS}}
    for var, (kind, payload) in actions.items()])

with pd.ExcelWriter(OUT_INTERMEDIATE, engine="openpyxl") as w:
    summary.to_excel(w, sheet_name="summary", index=False)
    template_long.to_excel(w, sheet_name="template_values", index=False)
    fe_table.to_excel(w, sheet_name="FE_ktoe", index=False)
    co2_by_tech.to_excel(w, sheet_name="CO2_by_tech_MtCO2", index=False)
    co2_by_tech.groupby("family")[YEARS].sum().reset_index().to_excel(
        w, sheet_name="CO2_family_split_MtCO2", index=False)
    seq_by_tech.to_excel(w, sheet_name="CCS_by_tech_MtCO2", index=False)
    coal_by_tech.to_excel(w, sheet_name="coal_by_tech_EJ", index=False)
    table(coef_log).to_excel(w, sheet_name="coal_coefficients", index=False)
    ist_nonco2[["GHG", "Units"] + YEARS].to_excel(w, sheet_name="nonCO2_Tg", index=False)
    prod_tech[["subsector", "technology", "Units"] + YEARS].to_excel(
        w, sheet_name="production_by_tech_Mt", index=False)
    table(exclusion_log).to_excel(w, sheet_name="excluded_inputs", index=False)
print(f"Saved {OUT_INTERMEDIATE}\nDone.")
